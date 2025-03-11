import logging
import os
import json
import re
import requests
import openpyxl
from docx import Document
from pathlib import Path
import yaml
from requests.exceptions import RequestException

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(Path(__file__).parent / 'partner_log.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def load_config(config_path='partner_config.yaml'):
    """加载配置文件"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)

        # 路径兼容性处理
        config['word_paths'] = [str(Path(p)) for p in config['word_paths']]
        config['excel_path'] = str(Path(config['excel_path']))

        logger.info("配置文件加载成功")
        return config
    except Exception as e:
        logger.error(f"配置文件加载失败: {str(e)}")
        raise


def read_word_file(file_path):
    """读取Word文档并清理格式 """
    try:
        doc = Document(file_path)
        text = []

        # 处理段落文本
        for para in doc.paragraphs:
            # 去除多余换行和特殊符号
            cleaned = re.sub(r'[\x00-\x1F\u3000\s]+', ' ', para.text).strip()
            if cleaned:
                text.append(cleaned)

        # 合并文本并去除目录符号
        full_text = '\n'.join(text)
        full_text = re.sub(r'\.{4,}', ' ', full_text)  # 去除长连字符
        full_text = re.sub(r'\n{3,}', '\n\n', full_text)  # 压缩多余空行

        logger.info(f"成功读取Word文件: {file_path}")
        return full_text
    except Exception as e:
        logger.error(f"读取Word文件失败: {file_path} - {str(e)}")
        raise


def process_excel_row(row, file_text, api_url, auth_token):
    """处理Excel单行数据"""
    try:
        # 构建Query参数
        query_data = {
            "条款": row[0],
            "内容概要填写说明和提示": row[1]
        }
        query_json = json.dumps(query_data, ensure_ascii=False, indent=4)

        # 构建请求体
        payload = {
            "response_mode": "blocking",
            "user": "haha-partner",
            "inputs": {
                "Query": query_json,
                "File": file_text
            }
        }

        headers = {'Authorization': auth_token}

        # 发送请求（设置10分钟超时）
        response = requests.post(
            api_url,
            json=payload,
            headers=headers,
            timeout=600
        )
        response.raise_for_status()

        # 解析响应数据
        response_data = response.json()
        # if response_data.get('event') != 'workflow_finished':
        #     raise ValueError("接口响应未完成")

        # 提取Answer中的JSON
        answer = response_data['data']['outputs']['Answer']
        json_match = re.search(r'```json(.*?)```', answer, re.DOTALL)
        if not json_match:
            raise ValueError("未找到有效JSON数据")

        result = json.loads(json_match.group(1).strip())
        return {
            '内容概要': convert_list_to_str(result.get('内容概要', '')),
            '条款序号': convert_list_to_str(result.get('条款序号', '')),
            '条款原文': convert_list_to_str(result.get('条款原文', '')),
            'error': ''
        }
    except Exception as e:
        logger.error(f"处理行失败: {row[0]} - {str(e)}")
        return {
            '内容概要': '处理失败',
            '条款序号': '处理失败',
            '条款原文': '处理失败',
            'error': str(e)
        }

def convert_list_to_str(data):
    """将列表转换为带换行符的字符串"""
    if isinstance(data, list):
        # 用换行符连接列表元素（Excel会识别为自动换行）
        return '\n'.join([str(item) for item in data])
    return str(data if data is not None else '未提取')


def process_documents(config):
    """主处理流程"""
    try:
        # 读取Excel模板
        wb = openpyxl.load_workbook(config['excel_path'])
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0]) + ['error']  # 添加错误列

        # 处理每个Word文件
        for word_path in config['word_paths']:
            try:
                # 读取Word内容
                file_text = read_word_file(word_path)

                # 创建新工作簿
                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.append(headers)

                # 处理每一行数据
                for idx, row in enumerate(rows[1:], start=1):
                    try:
                        result = process_excel_row(
                            row,
                            file_text,
                            config['api_url'],
                            config['auth_token']
                        )
                        new_row = list(row) + [result['error']]
                        new_row[2] = result['内容概要']
                        new_row[3] = result['条款序号']
                        new_row[4] = result['条款原文']
                        new_ws.append(new_row)
                    except Exception as e:
                        logger.error(f"第{idx}行处理异常: {str(e)}")
                        new_ws.append(list(row) + [str(e)])

                # 保存结果文件
                output_path = Path(word_path).with_name(
                    f"{Path(word_path).stem}_result.xlsx"
                )
                new_wb.save(output_path)
                logger.info(f"结果文件已保存: {output_path}")

            except Exception as e:
                logger.error(f"处理文件失败: {word_path} - {str(e)}")
                continue

    except Exception as e:
        logger.error(f"主流程异常: {str(e)}")
        raise


if __name__ == "__main__":
    try:
        config = load_config()
        process_documents(config)
    except Exception as e:
        logger.error(f"程序异常终止: {str(e)}")
